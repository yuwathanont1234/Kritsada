"""Build the Year-1 financial model as a polished xlsx with 3 scenarios."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

from model import (
    CONSERVATIVE, REALISTIC, OPTIMISTIC,
    run_month_by_month, summarize,
    TIER_PRICE_THB, TIER_SCAN_LIMIT, TIER_AVG_SCANS,
    COST_PER_SCAN_THB, FIXED_COSTS_THB, APP_STORE_TAKE_RATE,
)

# ── Styling helpers ─────────────────────────────────────────────────
GOLD = "ECC87A"
DARK = "1C1610"
LIGHT = "FFF8E7"
GRAY = "D9D9D9"
GREEN = "C6EFCE"
RED = "FFC7CE"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=DARK)
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="1C1610")
SUBHEADER_FILL = PatternFill("solid", fgColor=GOLD)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="9A7326")
SECTION_FILL = PatternFill("solid", fgColor=LIGHT)
BODY_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
NEGATIVE_FONT = Font(name="Calibri", size=10, color="C00000")

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")


def write_cover(ws):
    ws.title = "📋 Executive Summary"
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws['A1'] = "Luxury Authenticator — Year-1 Financial Projection"
    ws['A1'].font = Font(name="Calibri", size=18, bold=True, color="9A7326")
    ws.merge_cells('A1:D1')

    ws['A2'] = "งบประมาณการ ปีที่ 1 (รายเดือน) | 3 Scenarios"
    ws['A2'].font = Font(name="Calibri", size=12, italic=True, color="666666")
    ws.merge_cells('A2:D2')

    ws['A4'] = "Generated:"
    ws['B4'] = "2026-05-26"
    ws['A5'] = "Base unit:"
    ws['B5'] = "Thai Baht (THB)"
    ws['A6'] = "Exchange rate used:"
    ws['B6'] = "35 THB/USD"

    # Scenarios summary header
    row = 8
    ws.cell(row=row, column=1, value="KEY METRIC (สรุปรายปี)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    for c, label in enumerate(["Conservative (ต่ำ)", "Realistic (กลาง)", "Optimistic (สูง)"]):
        cell = ws.cell(row=row, column=c+2, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    scenarios = [run_month_by_month(s) for s in [CONSERVATIVE, REALISTIC, OPTIMISTIC]]
    last_months = [s[-1] for s in scenarios]

    rows_data = [
        ("Cumulative installs (M12)",
         [s[-1]['cumulative_installs'] for s in scenarios]),
        ("Active paid users (M12)",
         [s[-1]['active_paid_total'] for s in scenarios]),
        ("    • Standard ฿990",
         [s[-1]['active_paid_standard'] for s in scenarios]),
        ("    • Pro ฿1,990",
         [s[-1]['active_paid_pro'] for s in scenarios]),
        ("    • Premium ฿4,990",
         [s[-1]['active_paid_premium'] for s in scenarios]),
        ("Free users (cumulative)",
         [s[-1]['free_users'] for s in scenarios]),
        ("", []),
        ("Gross revenue (ปี 1)",
         [sum(r['gross_revenue'] for r in s) for s in scenarios]),
        ("App store fee (15%)",
         [sum(r['app_store_fee'] for r in s) for s in scenarios]),
        ("Net revenue (หลังหัก)",
         [sum(r['net_revenue'] for r in s) for s in scenarios]),
        ("", []),
        ("Variable costs (API scan)",
         [sum(r['variable_costs'] for r in s) for s in scenarios]),
        ("Fixed costs (infra)",
         [sum(r['fixed_costs'] for r in s) for s in scenarios]),
        ("Marketing spend",
         [sum(r['marketing'] for r in s) for s in scenarios]),
        ("Total costs",
         [sum(r['total_costs'] for r in s) for s in scenarios]),
        ("", []),
        ("Gross profit",
         [sum(r['gross_profit'] for r in s) for s in scenarios]),
        ("Gross margin %",
         [(sum(r['gross_profit'] for r in s) / sum(r['net_revenue'] for r in s) * 100) if sum(r['net_revenue'] for r in s) > 0 else 0 for s in scenarios]),
        ("NET PROFIT / LOSS (ปี 1)",
         [sum(r['net_profit'] for r in s) for s in scenarios]),
    ]

    for label, vals in rows_data:
        row += 1
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        if label.startswith("NET PROFIT") or label.startswith("Gross profit"):
            ws.cell(row=row, column=1).font = TOTAL_FONT
        if not vals:
            continue
        for c, v in enumerate(vals):
            cell = ws.cell(row=row, column=c+2)
            if "margin %" in label:
                cell.value = round(v, 1)
                cell.number_format = '0.0"%"'
            elif "users" in label.lower() or "installs" in label.lower():
                cell.value = v
                cell.number_format = '#,##0'
            else:
                cell.value = v
                cell.number_format = '#,##0;[Red]-#,##0'
            cell.alignment = RIGHT
            if "PROFIT" in label and v < 0:
                cell.font = Font(name="Calibri", size=10, color="C00000", bold=True)
            elif "PROFIT" in label:
                cell.font = Font(name="Calibri", size=10, color="2E7D32", bold=True)

    # Notes section
    row += 3
    ws.cell(row=row, column=1, value="ASSUMPTIONS").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    notes = [
        ("Cost per scan (steady state)", f"฿{COST_PER_SCAN_THB:.2f}"),
        ("  - Cache hit (60%)",          "฿0.30/scan"),
        ("  - Cheap-brand bypass (25%)", "฿1.10/scan"),
        ("  - Full pipeline (15%)",      "฿2.98/scan"),
        ("Fixed monthly (infra)",        f"฿{sum(FIXED_COSTS_THB.values()):,}/mo"),
        ("App store take rate",          "15% (Apple Small Business / Google Play)"),
        ("Std tier (฿990) avg scans/mo", f"{TIER_AVG_SCANS['standard']}"),
        ("Pro tier (฿1,990) avg scans",  f"{TIER_AVG_SCANS['pro']}"),
        ("Premium (฿4,990) avg scans",   f"{TIER_AVG_SCANS['premium']}"),
    ]
    for label, val in notes:
        row += 1
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        ws.cell(row=row, column=2, value=val).font = BODY_FONT


def write_scenario_sheet(ws, scenario):
    ws.title = scenario.label_th.replace(" (Base case)", "").replace("(", "").replace(")", "").strip()[:30]
    rows = run_month_by_month(scenario)

    # Column setup — A = label, B-M = M1..M12, N = Total
    ws.column_dimensions['A'].width = 32
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # Title
    ws['A1'] = f"{scenario.label_th} — Monthly P&L"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True, color="9A7326")
    ws.merge_cells('A1:N1')

    # Header row
    headers = ['METRIC'] + [f'M{i}' for i in range(1, 13)] + ['TOTAL Y1']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    section_groups = [
        ('━ ACQUISITION & USERS ━', [
            ('New installs',                  'new_installs',         '#,##0'),
            ('Cumulative installs',           'cumulative_installs',  '#,##0'),
            ('New paid conversions',          'new_paid',             '#,##0'),
            ('Churned paid',                  'churned_paid',         '#,##0'),
            ('Active — Standard ฿990',        'active_paid_standard', '#,##0'),
            ('Active — Pro ฿1,990',           'active_paid_pro',      '#,##0'),
            ('Active — Premium ฿4,990',       'active_paid_premium',  '#,##0'),
            ('Active paid (total)',           'active_paid_total',    '#,##0'),
            ('Free users (cumulative)',       'free_users',           '#,##0'),
        ]),
        ('━ REVENUE (THB) ━', [
            ('Revenue — Standard',            'rev_standard',         '#,##0'),
            ('Revenue — Pro',                 'rev_pro',              '#,##0'),
            ('Revenue — Premium',             'rev_premium',          '#,##0'),
            ('Gross revenue',                 'gross_revenue',        '#,##0'),
            ('App store fee (-15%)',          'app_store_fee',        '#,##0'),
            ('NET REVENUE',                   'net_revenue',          '#,##0'),
        ]),
        ('━ COSTS (THB) ━', [
            ('Free-user scan cost',           'free_scan_cost',       '#,##0'),
            ('Paid-user scan cost',           'paid_scan_cost',       '#,##0'),
            ('Variable costs (total)',        'variable_costs',       '#,##0'),
            ('Fixed costs (infra)',           'fixed_costs',          '#,##0'),
            ('Marketing spend',               'marketing',            '#,##0'),
            ('TOTAL COSTS',                   'total_costs',          '#,##0'),
        ]),
        ('━ BOTTOM LINE (THB) ━', [
            ('Gross profit',                  'gross_profit',         '#,##0'),
            ('Gross margin %',                'gross_margin_pct',     '0.0"%"'),
            ('NET PROFIT / (LOSS)',           'net_profit',           '#,##0;[Red](#,##0)'),
        ]),
    ]

    current_row = 4
    for group_label, fields in section_groups:
        ws.cell(row=current_row, column=1, value=group_label).font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=14)
        current_row += 1

        for label, key, num_fmt in fields:
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = TOTAL_FONT if label.isupper() else BODY_FONT
            cell.border = BORDER
            for m_idx, r in enumerate(rows):
                c = ws.cell(row=current_row, column=2 + m_idx, value=r[key])
                c.number_format = num_fmt
                c.alignment = RIGHT
                c.border = BORDER
                if key == 'net_profit' and r[key] < 0:
                    c.font = Font(name="Calibri", size=10, color="C00000", bold=True)
                elif key == 'net_profit':
                    c.font = Font(name="Calibri", size=10, color="2E7D32", bold=True)
            # Total column
            if key in ('cumulative_installs', 'active_paid_standard', 'active_paid_pro',
                      'active_paid_premium', 'active_paid_total', 'free_users',
                      'gross_margin_pct'):
                total = rows[-1][key]  # for "stocks" use final value
            else:
                total = sum(r[key] for r in rows)
            tcell = ws.cell(row=current_row, column=14, value=total)
            tcell.number_format = num_fmt
            tcell.alignment = RIGHT
            tcell.font = TOTAL_FONT
            tcell.fill = PatternFill("solid", fgColor="F0F0F0")
            tcell.border = BORDER
            current_row += 1
        current_row += 1  # spacer

    # Add chart for revenue / profit
    chart = LineChart()
    chart.title = "Monthly Net Revenue vs Net Profit"
    chart.style = 12
    chart.y_axis.title = 'THB'
    chart.x_axis.title = 'Month'

    # Find the row indices for net_revenue and net_profit
    # We'll re-find them by label name
    label_col = [(ws.cell(row=r, column=1).value, r) for r in range(4, current_row)]
    rev_row = next(r for l, r in label_col if l and 'NET REVENUE' in str(l))
    profit_row = next(r for l, r in label_col if l and 'NET PROFIT' in str(l))

    data_rev = Reference(ws, min_col=2, min_row=rev_row, max_col=13, max_row=rev_row)
    data_profit = Reference(ws, min_col=2, min_row=profit_row, max_col=13, max_row=profit_row)
    chart.add_data(data_rev, titles_from_data=False)
    chart.add_data(data_profit, titles_from_data=False)
    chart.series[0].tx = openpyxl_string("Net Revenue")
    chart.series[1].tx = openpyxl_string("Net Profit")
    cats = Reference(ws, min_col=2, min_row=3, max_col=13, max_row=3)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, f"A{current_row+2}")


def openpyxl_string(s):
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.data_source import StrRef
    sl = SeriesLabel()
    sl.v = s
    return sl


def write_assumptions(ws):
    ws.title = "🔧 Assumptions"
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 50

    ws['A1'] = "ASSUMPTIONS & MARKET RESEARCH BACKING"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True, color="9A7326")
    ws.merge_cells('A1:B1')

    sections = [
        ("Pricing (tier monthly, THB)", [
            ("Standard",  f"฿{TIER_PRICE_THB['standard']:,} → 50 scans/mo (already in app)"),
            ("Pro",       f"฿{TIER_PRICE_THB['pro']:,} → 100 scans/mo (already in app)"),
            ("Premium",   f"฿{TIER_PRICE_THB['premium']:,} → 200 scans/mo (already in app)"),
            ("Free",      "5 scans / 30 days, then locked"),
        ]),
        ("Cost per scan (post-optimization)", [
            ("Cache hit (60% of scans)",    "฿0.30 — price+identify cached"),
            ("Cheap-brand bypass (25%)",    "฿1.10 — auth bypassed, only identify+price"),
            ("Full pipeline (15%)",         "฿2.98 — full identify+auth+price+grounding"),
            ("Blended cost/scan",           f"฿{COST_PER_SCAN_THB:.2f}"),
        ]),
        ("Utilization (avg scans/user/mo)", [
            ("Free user",     f"{TIER_AVG_SCANS['free']} of 5 (70%)"),
            ("Standard",      f"{TIER_AVG_SCANS['standard']} of 50 (56%)"),
            ("Pro",           f"{TIER_AVG_SCANS['pro']} of 100 (55%)"),
            ("Premium",       f"{TIER_AVG_SCANS['premium']} of 200 (55% — dealer flow)"),
        ]),
        ("Fixed costs (THB/month)", [
            (k, f"฿{v}") for k, v in FIXED_COSTS_THB.items()
        ] + [("TOTAL", f"฿{sum(FIXED_COSTS_THB.values()):,}/mo")]),
        ("Conversion funnel benchmarks", [
            ("Install → Trial",  "18-32% (Conservative→Optimistic)"),
            ("Trial → Paid",     "15-27% (Conservative→Optimistic)"),
            ("SEA freemium avg", "1-3% install→paid (RevenueCat 2025)"),
            ("Our model",        "Effective install→paid ~3-9%"),
        ]),
        ("Market research sources (key)", [
            ("Thailand luxury watch market",   "$310M USD in 2025, 4% CAGR [Statista]"),
            ("Total watch market TH",          "$327M → $509M by 2034 [IMARC]"),
            ("Smartphone users TH (2026)",     "~60 million [Statista]"),
            ("Counterfeit prevalence",         "50% of fakes globally = Rolex [SCMP]"),
            ("Thailand counterfeit hotspots",  "MBK, Chinatown, Bang Niang — 809 watches seized 2025"),
            ("Entrupy (B2B competitor)",       "$139-$1,049/mo, 25-250 tokens"),
            ("LegitCheck (consumer)",          "$9.99/mo subscription + $10-20/auth"),
            ("eBay Auth Guarantee",            "Free for $2k+ watches, $80 below"),
            ("Chrono24 Certified",             "$249 per certification"),
            ("iOS ARPU global",                "$138 vs Android $72"),
        ]),
        ("Churn (monthly)", [
            ("Conservative",  "18% → declining to 10%"),
            ("Realistic",     "15% → declining to 8%"),
            ("Optimistic",    "12% → declining to 6%"),
            ("Industry SEA",  "8-15% typical for freemium"),
        ]),
        ("App store take rate", [
            ("Apple Small Biz Program",  "15% (first $1M revenue/year)"),
            ("Google Play",              "15% for first $1M/year"),
            ("Used in model",            f"{APP_STORE_TAKE_RATE*100:.0f}%"),
        ]),
    ]

    row = 3
    for title, items in sections:
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
        row += 1
        for label, val in items:
            ws.cell(row=row, column=1, value=label).font = BODY_FONT
            ws.cell(row=row, column=2, value=val).font = BODY_FONT
            row += 1
        row += 1


def write_tier_economics(ws):
    """Per-tier unit economics (LTV, payback, CAC)."""
    ws.title = "💎 Unit Economics"
    ws.column_dimensions['A'].width = 32
    for i in range(2, 6):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws['A1'] = "Per-Tier Unit Economics (Realistic scenario, blended)"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True, color="9A7326")
    ws.merge_cells('A1:E1')

    headers = ['METRIC', 'Free', 'Standard', 'Pro', 'Premium']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    tiers = ['free', 'standard', 'pro', 'premium']
    avg_scans = [TIER_AVG_SCANS[t] for t in tiers]
    prices = [0] + [TIER_PRICE_THB[t] for t in ['standard', 'pro', 'premium']]
    scan_costs = [a * COST_PER_SCAN_THB for a in avg_scans]
    revenues = [p for p in prices]
    net_rev = [p * (1 - APP_STORE_TAKE_RATE) for p in prices]
    margins = [(nr - sc) for nr, sc in zip(net_rev, scan_costs)]
    margin_pcts = [(m / nr * 100) if nr > 0 else 0 for m, nr in zip(margins, net_rev)]

    rows = [
        ('Monthly price (฿)',          prices,        '#,##0'),
        ('Avg scans/mo',               avg_scans,     '0.0'),
        ('Scan cost (฿)',              scan_costs,    '#,##0.00'),
        ('Gross revenue (฿)',          revenues,      '#,##0'),
        ('Net rev after app store (฿)', net_rev,      '#,##0'),
        ('Net margin (฿)',             margins,       '#,##0'),
        ('Net margin %',               margin_pcts,   '0.0"%"'),
    ]

    row = 4
    for label, vals, fmt in rows:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        for c, v in enumerate(vals):
            cell = ws.cell(row=row, column=c+2, value=v)
            cell.number_format = fmt
            cell.alignment = RIGHT
        row += 1

    # LTV section
    row += 2
    ws.cell(row=row, column=1, value="LIFETIME VALUE (LTV)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    avg_churn = 0.10  # realistic blended
    avg_lifetime_mo = 1 / avg_churn  # 10 months average

    ws.cell(row=row, column=1, value="Assumed churn (monthly)").font = BODY_FONT
    ws.cell(row=row, column=2, value=f"{avg_churn*100:.0f}%")
    row += 1
    ws.cell(row=row, column=1, value="Avg lifetime (months)").font = BODY_FONT
    ws.cell(row=row, column=2, value=f"{avg_lifetime_mo:.1f}")
    row += 2

    for i, t in enumerate(['standard', 'pro', 'premium']):
        ltv = margins[i+1] * avg_lifetime_mo
        ws.cell(row=row, column=1, value=f"LTV — {t.title()}").font = TOTAL_FONT
        ws.cell(row=row, column=2, value=ltv).number_format = '#,##0'
        ws.cell(row=row, column=2).font = TOTAL_FONT
        row += 1


# ── Main ─────────────────────────────────────────────────────────────
wb = Workbook()
write_cover(wb.active)
write_scenario_sheet(wb.create_sheet(), CONSERVATIVE)
write_scenario_sheet(wb.create_sheet(), REALISTIC)
write_scenario_sheet(wb.create_sheet(), OPTIMISTIC)
write_assumptions(wb.create_sheet())
write_tier_economics(wb.create_sheet())

output = '/Users/kritsada/Desktop/Luxury-authenticator/financial_model/Luxury_Authenticator_Year1_Projection.xlsx'
wb.save(output)
print(f"✅ Saved: {output}")
import os
print(f"   Size: {os.path.getsize(output)/1024:.1f} KB")
